import tkinter as tk
from tkinter import messagebox
import win32com.client as win32
import os
import time
from pdf2image import convert_from_path
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Font, Border, Side
from openpyxl.utils import get_column_letter
import datetime
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.image import MIMEImage
from email.mime.application import MIMEApplication
import openpyxl
import pythoncom
import logging
import ssl

logging.basicConfig(
    filename='email_errors.log',
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)

# Configuração do servidor SMTP
email_remetente = os.getenv('EMAIL_REMETENTE')
nome_usuario_smtp = os.getenv('USUARIO_SMTP')
senha_smtp = os.getenv('SENHA_SMTP')
servidor_smtp = os.getenv('SERVIDOR_SMTP')
porta_smtp = int(os.getenv('PORTA_SMTP', 587))

file_path = os.getenv('CAMINHO_ARQUIVO_EXCEL')
directory = os.getenv('CAMINHO_PASTA_PARCEIROS')
poppler_path = os.getenv('CAMINHO_POPPLER')
macro_name = 'ExportPDFsFromDropDown'

# Validação simples para garantir que o .env foi lido
if not all([email_remetente, senha_smtp, file_path]):
    messagebox.showerror("Erro de Configuração", "Variáveis de ambiente não encontradas. Verifique o arquivo .env")
    exit()
# ------------------------------------------

def atualizar_base():
    aguarde_janela = tk.Toplevel(root)
    aguarde_janela.title("Processando")
    aguarde_label = tk.Label(aguarde_janela, text="Aguarde, estamos processando...")
    aguarde_label.pack(padx=30, pady=20)
    aguarde_janela.update()

    try:
        open_excel_and_update(file_path, macro_name, directory, poppler_path)
        messagebox.showinfo("Concluído", "Arquivo atualizado com sucesso.")
    except Exception as e:
        messagebox.showerror("Erro", f"Ocorreu um erro: {str(e)}")
    finally:
        aguarde_janela.destroy()

def open_excel_and_update(file_path, macro_name, directory, poppler_path):
    excel = None
    wb = None
    try:
        pythoncom.CoInitialize()
        
        # Configuração robusta do Excel
        excel = win32.DispatchEx('Excel.Application')
        excel.Visible = False
        excel.DisplayAlerts = False
        excel.AskToUpdateLinks = False
        excel.EnableEvents = False
        
        # Abre o workbook - FORMA CORRETA SEM PARÂMETROS NOMEADOS
        wb = excel.Workbooks.Open(file_path, False, False, None, None, None, True, 1)
        print(f"Arquivo Excel {file_path} aberto com sucesso.")

        # Atualização de conexões com tratamento aprimorado
        print("Atualizando conexões de dados...")
        for i in range(wb.Connections.Count):
            try:
                conn = wb.Connections.Item(i+1)  # Índices COM começam em 1
                if conn.Type == 1:  # xlConnectionTypeOLEDB
                    conn.OLEDBConnection.BackgroundQuery = False
                conn.Refresh()
                print(f"Conexão {conn.Name} atualizada.")
                time.sleep(1)
            except Exception as e:
                print(f"Erro na conexão {i+1}: {str(e)}")
                continue

        # Espera ativa para conclusão
        start_time = time.time()
        while time.time() - start_time < 300:  # 5 minutos máximo
            refreshing = False
            try:
                for conn in wb.Connections:
                    if conn.Refreshing:
                        refreshing = True
                        break
                if not refreshing:
                    break
                time.sleep(5)
            except:
                break

        # Execução da macro com tratamento especial
        print(f"Executando a macro {macro_name}...")
        try:
            excel.Application.Run("ExportPDFsFromDropDown")
            time.sleep(2)
            excel.Application.Run(f"'{wb.Name}'!{macro_name}")
            
            for _ in range(30):
                try:
                    if excel.Ready:
                        break
                    time.sleep(1)
                except:
                    break
        except Exception as e:
            print(f"Erro na execução da macro: {str(e)}")
            raise

        try:
            wb.Save()
            print("Workbook salvo com sucesso.")
        except Exception as e:
            print(f"Erro ao salvar: {str(e)}")
            try:
                wb.SaveAs(file_path)
            except Exception as e2:
                print(f"Falha crítica ao salvar: {str(e2)}")
                raise

    except Exception as e:
        print(f"Erro crítico: {str(e)}")
        raise
    finally:
        # Liberação de recursos em ordem inversa
        try:
            if 'wb' in locals() and wb is not None:
                wb.Close(True)
        except Exception as e:
            print(f"Erro ao fechar workbook: {str(e)}")

        try:
            if 'excel' in locals() and excel is not None:
                excel.EnableEvents = True
                excel.Quit()
        except Exception as e:
            print(f"Erro ao fechar Excel: {str(e)}")

        pythoncom.CoUninitialize()
        del wb
        del excel
        time.sleep(2)
        print("Processo finalizado.")

    # Função para converter PDFs para imagens
    convert_pdfs_to_images(directory, poppler_path)

def convert_pdfs_to_images(directory, poppler_path):
    for filename in os.listdir(directory):
        if filename.endswith(".pdf"):
            pdf_path = os.path.join(directory, filename)
            try:
                images = convert_from_path(pdf_path, poppler_path=poppler_path)
                for i, image in enumerate(images):
                    image_filename = f"{os.path.splitext(filename)[0]}.png"
                    image_path = os.path.join(directory, image_filename)
                    image.save(image_path, 'PNG')
                os.remove(pdf_path)
                print(f"{filename} convertido e excluído.")
            except Exception as e:
                print(f"Erro ao processar {filename}: {str(e)}")

def gerar_tabelas():
    aguarde_janela = tk.Toplevel(root)
    aguarde_janela.title("Processando")
    tk.Label(aguarde_janela, text="Aguarde, estamos processando...").pack(padx=20, pady=20)
    aguarde_janela.update()  # Força a atualização da janela de status
    print("Gerando tabelas...")

    def limpar_nome_arquivo(nome):
        if nome is None:
            return 'Nome_Indefinido'
        invalid_chars = '<>:"/\\|?*'
        for char in invalid_chars:
            nome = nome.replace(char, '')
        return nome

    def aplicar_estilos(ws):
        fill = PatternFill(start_color='D63066', end_color='D63066', fill_type='solid')
        font = Font(color='FFFFFF')
        thin = Side(border_style="thin", color="000000")
        border = Border(top=thin, left=thin, right=thin, bottom=thin)
        
        for row in ws.iter_rows(min_row=1, max_row=1):
            for cell in row:
                cell.fill = fill
                cell.font = font
        
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.border = border

        for col in range(1, ws.max_column + 1):
            ws.column_dimensions[get_column_letter(col)].width = 15

    arquivo = "C:\\Parceiros\\Parceiros.xlsm"
    diretorio = "C:\\Parceiros\\"

    data_hoje = datetime.datetime.now().date()
    if data_hoje.day == 1:
        primeiro_dia_mes = pd.Timestamp((data_hoje.replace(day=1) - datetime.timedelta(days=1)).replace(day=1))
        ultimo_dia_mes = pd.Timestamp(data_hoje.replace(day=1) - datetime.timedelta(days=1))
    else:
        primeiro_dia_mes = pd.Timestamp(data_hoje.replace(day=1))
        ultimo_dia_mes = pd.Timestamp((data_hoje.replace(day=28) + datetime.timedelta(days=4)).replace(day=1) - datetime.timedelta(days=1))

    wb = load_workbook(arquivo, data_only=True)
    abas = ['VENDAS', 'INSTALAÇÕES', 'CANCELAMENTO', 'CHURN', 'MOVEL', 'SUSPENSOS', 'SAFRA']

    df_parceiros = pd.DataFrame(wb['Parceiro'].values)
    df_parceiros.columns = df_parceiros.iloc[0]
    df_parceiros = df_parceiros[1:]
    lista_de_parceiros = df_parceiros['CANAL_VENDAS'].unique()

    renomear_dia = {
        'VENDAS': 'DATA_CADASTRO',
        'INSTALAÇÕES': 'DATA_HABILITAÇÃO',
        'CANCELAMENTO': 'DATA_CANCELAMENTO',
        'CHURN': 'DATA_CANCELAMENTO',
        'MOVEL' : 'DATA_VENDA'
    }

    for aba in abas:
        print(f"Processando a aba: {aba}")
        df = pd.DataFrame(wb[aba].values)
        df.columns = df.iloc[0]
        df = df[1:]
        if aba not in ['SUSPENSOS', 'SAFRA', 'MIGRACAO']:
            df['Dia'] = pd.to_datetime(df['DIA'], errors='coerce').dt.strftime('%d/%m/%Y')
            df.rename(columns={'Dia': renomear_dia[aba]}, inplace=True)
            df = df[(pd.to_datetime(df[renomear_dia[aba]], format='%d/%m/%Y') >= primeiro_dia_mes) & (pd.to_datetime(df[renomear_dia[aba]], format='%d/%m/%Y') <= ultimo_dia_mes)]
        
        if aba in ['INSTALAÇÕES', 'CANCELAMENTO', 'CHURN']:
            df['DATA_CADASTRO'] = pd.to_datetime(df['DATA_CADASTRO']).dt.strftime('%d/%m/%Y')
        
        if aba == 'VENDAS':
            colunas_desejadas = [renomear_dia[aba], 'ID_CONTRATO', 'COD_CLIENTE', 'NOME_PLANO_ATUAL', 'CIDADE_HIERARQUIA', 'REGIONAL','MACRO_REGIAO', 'STATUS_CONTRATO', 'CANAL_VENDAS']
        elif aba == 'INSTALAÇÕES':
            colunas_desejadas = [renomear_dia[aba], 'DATA_CADASTRO', 'ID_CONTRATO', 'COD_CLIENTE', 'NOME_PLANO_ATUAL','CIDADE_HIERARQUIA', 'REGIONAL', 'MACRO_REGIAO', 'STATUS_CONTRATO', 'CANAL_VENDAS', 'VALOR_CONTRATO']
        elif aba  == 'CHURN':
            colunas_desejadas = [renomear_dia[aba], 'DATA_CADASTRO', 'ID_CONTRATO', 'COD_CLIENTE', 'NOME_PLANO_ATUAL', 'CIDADE_HIERARQUIA', 'REGIONAL', 'MACRO_REGIAO', 'STATUS_CONTRATO', 'CANAL_VENDAS', 'MOTIVO_CANCELAMENTO']
        elif aba == 'CANCELAMENTO':
            colunas_desejadas = [renomear_dia[aba], 'DATA_CADASTRO', 'ID_CONTRATO', 'COD_CLIENTE', 'NOME_PLANO_ATUAL', 'CIDADE_HIERARQUIA', 'REGIONAL', 'MACRO_REGIAO', 'STATUS_CONTRATO', 'CANAL_VENDAS', 'TIPO_CANCELAMENTO', 'MOTIVO_CANCELAMENTO']
        elif aba == 'MOVEL':
            if 'DIA' in df.columns:
                df['DATA_VENDA'] = pd.to_datetime(df['DIA'], errors='coerce').dt.strftime('%d/%m/%Y')
                df.drop(columns=['DIA'], inplace=True)
            for coluna_data in ['DAT_ENVIO', 'DAT_ENTREGA']:
                if coluna_data in df.columns:
                    df[coluna_data] = pd.to_datetime(df[coluna_data], errors='coerce').dt.strftime('%d/%m/%Y')
            
            colunas_desejadas = ['DATA_VENDA'] + [col for col in df.columns if col != 'DATA_VENDA']
        elif aba == 'SUSPENSOS':
            if 'DIA' in df.columns:
                df['DIA'] = pd.to_datetime(df['DIA'], errors='coerce').dt.strftime('%d/%m/%Y')
            colunas_desejadas = df.columns.tolist()

        elif aba == 'SAFRA':
            for coluna in ['DATA_CADASTRO', 'VENCIMENTO', 'PAGAMENTO']:
                if coluna in df.columns:
                    df[coluna] = (pd.to_datetime(df[coluna], errors='coerce').dt.strftime('%d/%m/%Y'))
            colunas_desejadas = df.columns.tolist()

        """elif aba == 'MIGRACAO':
            if 'DATA_HABILITACAO' in df.columns:
                df['DATA_HABILITACAO'] = pd.to_datetime(df['DATA_HABILITACAO'], errors='coerce').dt.strftime('%d/%m/%Y')
            colunas_desejadas = df.columns.tolist() """


        for parceiro in lista_de_parceiros:
            df_filtrado = df[df['CANAL_VENDAS'] == parceiro][colunas_desejadas]
            df_filtrado = df_filtrado.sort_values(by=renomear_dia[aba]) if aba not in ['SUSPENSOS', 'SAFRA', 'MIGRACAO'] else df_filtrado
            nome_arquivo = limpar_nome_arquivo(parceiro) + f'_{aba}.xlsx'
            caminho_completo = os.path.join(diretorio, nome_arquivo)
            
            writer = pd.ExcelWriter(caminho_completo, engine='openpyxl')
            df_filtrado.to_excel(writer, index=False)
            ws = writer.sheets['Sheet1']
            aplicar_estilos(ws)
            writer.close()
            
            print(f'Arquivo criado para {aba}: {caminho_completo}')

    for parceiro in lista_de_parceiros:
        with pd.ExcelWriter(os.path.join(diretorio, limpar_nome_arquivo(parceiro) + '.xlsx'), engine='openpyxl') as writer:
            for aba in abas:
                nome_arquivo = limpar_nome_arquivo(parceiro) + f'_{aba}.xlsx'
                caminho_completo = os.path.join(diretorio, nome_arquivo)
                if os.path.exists(caminho_completo):
                    df = pd.read_excel(caminho_completo)
                    df.to_excel(writer, sheet_name=aba, index=False)
                    ws = writer.sheets[aba]
                    aplicar_estilos(ws)

                    os.remove(caminho_completo)
                    print(f'Arquivo temporário excluído: {caminho_completo}')

    aguarde_janela.destroy()
    messagebox.showinfo("Concluído", "Tabelas geradas com sucesso.")
    print("Todos os processos foram concluídos com sucesso!")


# Funções adicionais para envio de e-mails
def ler_planilha(caminho_planilha):
    wb = openpyxl.load_workbook(caminho_planilha)
    sheet = wb['Parceiro']
    destinatarios = []

    for row in sheet.iter_rows(min_row=2, values_only=True):
        canal_vendas = row[0]
        
        # E-mail principal (coluna 2) agora tratado em lista:
        email = []
        if row[1]:
            raw = row[1]
            if isinstance(raw, str):
                email = [
                    addr.strip().lower()
                    for addr in raw.split(';')
                    if addr.strip()
                ]
            else:
                addr = str(raw).strip().lower()
                if addr:
                    email = [addr]

        # E-mails em cópia (coluna 3)
        emails_copia = []
        if row[2]:
            raw_cc = row[2]
            if isinstance(raw_cc, str):
                emails_copia = [
                    addr.strip().lower()
                    for addr in raw_cc.split(';')
                    if addr.strip()
                ]
            else:
                addr = str(raw_cc).strip().lower()
                if addr:
                    emails_copia = [addr]

        nome_empresa = row[3] if len(row) > 3 else ""

        # Só adiciona se existir canal e pelo menos um e-mail principal
        if canal_vendas and email:
            destinatarios.append((canal_vendas, email, emails_copia, nome_empresa))

    return destinatarios

def encontrar_arquivo(caminho_pasta, nome_arquivo, extensao):
    nome_arquivo_com_extensao = f"{nome_arquivo}{extensao}"
    for arquivo in os.listdir(caminho_pasta):
        if arquivo.lower() == nome_arquivo_com_extensao.lower():
            return os.path.join(caminho_pasta, arquivo)
    return None

def verificar_conexao(servidor):
    """Verifica se a conexão SMTP está ativa."""
    try:
        status = servidor.noop()[0]
        return True if status == 250 else False
    except Exception as e:
        logging.error(f"Falha na verificação de conexão: {str(e)}")
        return False

def reconectar_servidor(ctx):
    """Reconecta ao servidor SMTP com tratamento de erros."""
    max_tentativas = 7
    for tentativa in range(max_tentativas):
        try:
            servidor = smtplib.SMTP(servidor_smtp, porta_smtp, timeout=60)
            servidor.sock.settimeout(120)
            servidor.set_debuglevel(1)
            servidor.ehlo()
            servidor.starttls(context=ctx)
            servidor.ehlo()
            servidor.login(nome_usuario_smtp, senha_smtp)
            logging.info("Conexão SMTP reestabelecida com sucesso")
            return servidor
        except Exception as e:
            logging.error(f"Tentativa {tentativa+1} de reconexão falhou: {str(e)}")
            if tentativa == max_tentativas - 1:
                raise
            time.sleep(10)
    return None

def enviar_emails():
    aguarde_janela = tk.Toplevel(root)
    aguarde_janela.title("Processando")

    progresso_frame = tk.Frame(aguarde_janela)
    progresso_frame.pack(padx=20, pady=10)
    
    tk.Label(progresso_frame, text="Aguarde, estamos processando...").pack()
    contador_label = tk.Label(progresso_frame, text="0/0")
    contador_label.pack()
    
    aguarde_janela.update()

    try:
        caminho_planilha = file_path
        pasta_arquivos = directory
        destinatarios = ler_planilha(caminho_planilha)
        total = len(destinatarios)
        
        ctx = ssl.create_default_context()
        ctx.minimum_version = ssl.TLSVersion.TLSv1_2  
        
        servidor = None
        emails_enviados = 0
        max_emails_por_conexao = 10  # Ajuste conforme necessidade do servidor

        for index, (nome, email, emails_copia, nome_empresa) in enumerate(destinatarios, 1):
            contador_label.config(text=f"{index}/{total}")
            aguarde_janela.update()
            
            # Verifica necessidade de reconexão
            if servidor is None or emails_enviados >= max_emails_por_conexao or not verificar_conexao(servidor):
                if servidor:
                    try:
                        servidor.quit()  # Tenta fechar a conexão se estiver ativa
                    except Exception as e:
                        logging.warning(f"Erro ao fechar conexão anterior: {str(e)}")
                servidor = reconectar_servidor(ctx)
                emails_enviados = 0

            # Tenta envio com 3 tentativas
            for tentativa in range(3):
                try:
                    caminho_excel = encontrar_arquivo(pasta_arquivos, nome, '.xlsx')
                    caminho_imagem = encontrar_arquivo(pasta_arquivos, nome, '.png')
                    
                    enviar_email(
                        servidor,
                        email,
                        emails_copia,
                        caminho_excel,
                        caminho_imagem,
                        f'Relatório de Vendas - {nome}',
                        nome,
                        nome_empresa
                    )
                    emails_enviados += 1
                    logging.info(f"Sucesso: {', '.join(email)}")
                    break
                except Exception as e:
                    logging.error(f"Tentativa {tentativa+1} para {', '.join(email)}: {str(e)}")
                    if tentativa == 2:
                        #messagebox.showwarning("Erro", f"Falha ao enviar para {', '.join(email)}")
                        pass 
                    time.sleep(5)
            
            time.sleep(10)  # Intervalo maior entre envios

    except Exception as e:
        logging.critical(f"Erro crítico: {str(e)}")
        messagebox.showerror("Erro Grave", f"Falha no processo: {str(e)}")
    finally:
        if servidor:
            try:
                servidor.quit()  # Tenta fechar a conexão de forma segura
            except Exception as e:
                logging.warning(f"Erro ao fechar a conexão SMTP: {str(e)}")
        aguarde_janela.destroy()
        messagebox.showinfo("Concluído", "Processo de envio finalizado!")



def enviar_email(servidor, destinatarios, emails_copia, caminho_arquivo, imagem, texto_assunto, nome_parceiro, nome_empresa):
    """Função aprimorada com validações e tratamento de erros."""
    try:
        # Validação de arquivos
        anexos = []
        if caminho_arquivo and os.path.exists(caminho_arquivo):
            with open(caminho_arquivo, 'rb') as arquivo:
                part = MIMEApplication(arquivo.read(), Name=os.path.basename(caminho_arquivo))
                part['Content-Disposition'] = f'attachment; filename="{os.path.basename(caminho_arquivo)}"'
                anexos.append(part)
        else:
            logging.warning(f"Anexo não encontrado para {nome_parceiro}: {caminho_arquivo}")

        if imagem and os.path.exists(imagem):
            with open(imagem, 'rb') as img:
                msg_image = MIMEImage(img.read())
                msg_image.add_header('Content-ID', '<image1>')
                anexos.append(msg_image)
        else:
            logging.warning(f"Imagem não encontrada para {nome_parceiro}: {imagem}")

        # Construção da mensagem
        msg = MIMEMultipart('related')
        msg['From'] = email_remetente
        msg['To'] = ", ".join(destinatarios)
        if emails_copia:
            msg['Cc'] = ", ".join(emails_copia)
        msg['Subject'] = texto_assunto

        html = f"""
        <html>
        <body>
            <p>Olá {nome_empresa}!<br>Aqui está o seu relatório diário.</p>
            
            <div style="font-size: 14px;">
                <strong style="font-size: 16px;">OBSERVAÇÃO:</strong><br>
                * Esse email é de <strong style="font-size: 16px;">uso exclusivo para envio dos relatórios</strong>, qualquer dúvida acionar o seu gestor.<br>
                ** Este relatório oferece apenas uma visão diária <strong style="font-size: 16px;">de realizado para fins de conferência e acompanhamento</strong>, e não impacta nas regras de comissionamento. É importante considerar os dados <strong style="font-size: 16px;">enviados pelo time de RVV para o fechamento oficial</strong>.
            </div>
            
            <!-- Espaço adicional inserido aqui -->
            <br>
            <img src="cid:image1" alt="Relatório diário">
        </body>
        </html>
                """
        corpo = MIMEText(html, 'html')
        msg.attach(corpo)

        for anexo in anexos:
            msg.attach(anexo)

        # Envio seguro
        servidor.send_message(msg)
        logging.info(f"E-mail para {', '.join(destinatarios)} enviado com sucesso")

    except Exception as e:
        logging.error(f"Falha ao enviar para {', '.join(destinatarios)}: {str(e)}")
        raise

root = tk.Tk()
root.title("Relatório diário Parceiros")
root.geometry('400x200')  

button_frame = tk.Frame(root)
button_frame.pack(pady=20)

btn_atualizar_base = tk.Button(button_frame, text="Atualizar base", command=atualizar_base, width=15)
btn_gerar_tabelas = tk.Button(button_frame, text="Gerar tabelas", command=gerar_tabelas, width=15)
btn_enviar_emails = tk.Button(button_frame, text="Enviar e-mails", command=enviar_emails, width=15)

btn_atualizar_base.pack(side=tk.TOP, fill=tk.X, padx=20, pady=10)
btn_gerar_tabelas.pack(side=tk.TOP, fill=tk.X, padx=20, pady=10)
btn_enviar_emails.pack(side=tk.TOP, fill=tk.X, padx=20, pady=10)

# Inicia a execução da janela
root.mainloop()